from pathlib import Path
import random

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
RANGE_FILE = ROOT / "range_test_sample.xlsx"
VALUE_FILE = ROOT / "value_test_sample.xlsx"
SYSTEM_FILE = ROOT / "system_test_sample.xlsx"
OUTPUT_FILE = ROOT / "integrated_test_sample.xlsx"
DATA_LENGTH = 8
COLUMN_COUNT = 21


def read_sheet_rows(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        values = ["" if cell is None else str(cell).strip() for cell in row]
        if any(values):
            rows.append(values)
    return rows


def header_map(header):
    result = {}
    for index, name in enumerate(header):
        key = str(name).strip().upper().replace(" ", "").replace("_", "")
        if key == "INDEX(HEX)":
            key = "INDEX"
        if key == "RETURNVALUE":
            result["RETURN"] = index
        if key == "EXPECTEDVALUE":
            result["EXPECTED"] = index
        result[key] = index
    return result


def value_at(row, mapping, name):
    key = name.strip().upper().replace(" ", "").replace("_", "")
    index = mapping.get(key)
    if index is None or index >= len(row):
        return ""
    return row[index]


def make_empty_row(test_type):
    row = [""] * COLUMN_COUNT
    row[0] = test_type
    return row


def load_range_rows():
    rows = read_sheet_rows(RANGE_FILE)
    if not rows:
        return []
    mapping = header_map(rows[0])
    out = []
    for source in rows[1:]:
        row = make_empty_row("RANGE_TEST")
        row[1] = value_at(source, mapping, "INDEX")
        row[2] = value_at(source, mapping, "Description")
        row[3] = value_at(source, mapping, "MIN")
        row[4] = value_at(source, mapping, "MAX")
        row[5] = value_at(source, mapping, "DELAY ms")
        row[6] = value_at(source, mapping, "RETURN")
        row[7] = value_at(source, mapping, "PASS")
        row[8] = value_at(source, mapping, "FAIL")
        row[9] = value_at(source, mapping, "TOTAL")
        out.append(row)
    return out


def load_value_rows():
    rows = read_sheet_rows(VALUE_FILE)
    if not rows:
        return []
    mapping = header_map(rows[0])
    out = []
    for source in rows[1:]:
        row = make_empty_row("VALUE_TEST")
        row[1] = value_at(source, mapping, "HDR")
        row[2] = value_at(source, mapping, "LENGTH")
        row[3] = value_at(source, mapping, "TYPE")
        row[4] = value_at(source, mapping, "CATEGORY")
        row[5] = value_at(source, mapping, "OPCODE")
        for index in range(DATA_LENGTH):
            row[6 + index] = value_at(source, mapping, f"DATA[{index}]")
        row[14] = value_at(source, mapping, "DELAY ms")
        row[15] = value_at(source, mapping, "EXPECTED")
        row[16] = value_at(source, mapping, "RETURN")
        row[17] = value_at(source, mapping, "PASS")
        row[18] = value_at(source, mapping, "FAIL")
        row[19] = value_at(source, mapping, "TOTAL")
        out.append(row)
    return out


def load_system_rows():
    rows = read_sheet_rows(SYSTEM_FILE)
    if not rows:
        return []
    mapping = header_map(rows[0])
    out = []
    for source in rows[1:]:
        row = make_empty_row("SYS_TEST")
        row[1] = value_at(source, mapping, "TYPE")
        row[2] = value_at(source, mapping, "EXPECTED")
        row[3] = value_at(source, mapping, "RETURN")
        out.append(row)
    return out


def write_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "IntegratedTest"

    headers = [[""] * COLUMN_COUNT for _ in range(3)]
    headers[0][0:8] = ["SYS TEST", "TYPE", "EXPECTED", "RETURN", "PASS", "FAIL", "TOTAL", "EXCUTION"]
    headers[1][0:11] = ["RANGE TEST", "INDEX", "Description", "MIN", "MAX", "DELAY ms", "RETURN", "PASS", "FAIL", "TOTAL", "EXCUTION"]
    headers[2][0] = "VALUE TEST"
    headers[2][1:6] = ["HDR", "LENGTH", "TYPE", "CATEGORY", "OPCODE"]
    for index in range(DATA_LENGTH):
        headers[2][6 + index] = f"D[{index}]"
    headers[2][14:21] = ["DELAY ms", "EXPECTED", "RETURN", "PASS", "FAIL", "TOTAL", "EXCUTION"]

    for values in headers + rows:
        ws.append(values)

    dark_gray = PatternFill("solid", fgColor="7F7F7F")
    light_gray = PatternFill("solid", fgColor="D9D9D9")
    dark_green = PatternFill("solid", fgColor="80B080")
    light_green = PatternFill("solid", fgColor="F1F7ED")
    dark_yellow = PatternFill("solid", fgColor="BF8F00")
    light_yellow = PatternFill("solid", fgColor="FFF2CC")
    white_font = Font(color="FFFFFF")
    black_font = Font(color="000000")
    pass_font = Font(color="003366")
    fail_font = Font(color="990000")

    theme = {
        "SYS_TEST": (dark_gray, light_gray),
        "SYS TEST": (dark_gray, light_gray),
        "S": (dark_gray, light_gray),
        "RANGE_TEST": (dark_green, light_green),
        "RANGE TEST": (dark_green, light_green),
        "R": (dark_green, light_green),
        "VALUE_TEST": (dark_yellow, light_yellow),
        "VALUE TEST": (dark_yellow, light_yellow),
        "V": (dark_yellow, light_yellow),
    }

    for row_index in range(1, ws.max_row + 1):
        test_type = ws.cell(row=row_index, column=1).value
        dark_fill, light_fill = theme.get(test_type, (dark_gray, light_gray))
        is_header = row_index <= 3
        for col_index in range(1, COLUMN_COUNT + 1):
            cell = ws.cell(row=row_index, column=col_index)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.value in (None, ""):
                continue
            if is_header or col_index == 1:
                cell.fill = dark_fill
                cell.font = white_font
            else:
                cell.fill = light_fill
                cell.font = black_font
            if str(cell.value).upper() == "PASS":
                cell.font = pass_font
            elif str(cell.value).upper() == "FAIL":
                cell.font = fail_font

    for col_index in range(1, COLUMN_COUNT + 1):
        max_len = 10
        for row_index in range(1, ws.max_row + 1):
            value = ws.cell(row=row_index, column=col_index).value
            if value not in (None, ""):
                max_len = max(max_len, len(str(value)) + 2)
        ws.column_dimensions[get_column_letter(col_index)].width = min(max_len, 24)

    ws.freeze_panes = "A4"
    wb.save(OUTPUT_FILE)


def main():
    rows = load_range_rows() + load_value_rows() + load_system_rows()
    random.Random(0x1234ABCD).shuffle(rows)
    write_workbook(rows)
    print(f"created {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
